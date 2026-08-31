"""Streaming xlsx writer (openpyxl write_only mode) used by the export engine.

Writes a Metadata sheet first (BRD §6.6: exports retain filters and metadata and
carry their classification), then the data sheet. Writes to a temp file and
atomically replaces the target, so concurrent lazy regeneration is safe.
"""

import json
import os
import uuid
from collections.abc import AsyncIterator
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

_CELL_CHUNK = 500


def _cell(value: Any) -> Any:
    """Coerce a model value into an xlsx-safe cell value."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, uuid.UUID):
        return str(value)
    if isinstance(value, dict | list):
        return json.dumps(value, default=str)
    return value


async def write_export_xlsx(
    path: Path,
    *,
    metadata: dict[str, Any],
    metadata_sheet: str,
    data_sheet: str,
    headers: list[str],
    rows: AsyncIterator[list[Any]],
) -> int:
    """Stream ``rows`` into ``path``; return the number of data rows written."""
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = path.with_name(f"{path.name}.{uuid.uuid4().hex[:8]}.tmp")

    workbook = Workbook(write_only=True)
    meta_ws = workbook.create_sheet(title=metadata_sheet)
    meta_ws.append(["Field", "Value"])
    for key, value in metadata.items():
        meta_ws.append([key, _cell(value)])

    data_ws = workbook.create_sheet(title=data_sheet)
    data_ws.append(headers)

    count = 0
    buffer: list[list[Any]] = []
    async for row in rows:
        buffer.append([_cell(value) for value in row])
        count += 1
        if len(buffer) >= _CELL_CHUNK:
            for buffered in buffer:
                data_ws.append(buffered)
            buffer.clear()
    for buffered in buffer:
        data_ws.append(buffered)

    workbook.save(tmp_path)
    os.replace(tmp_path, path)
    return count
