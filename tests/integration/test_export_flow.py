"""Export integration tests: real sqlite DB, real file generation, full lifecycle.

These exercise the engine against an actual database (unlike the unit tests,
which mock all DB access). Tables are created for the test run only.
"""

import uuid
from datetime import UTC, datetime, timedelta
from pathlib import Path

import pytest
from openpyxl import load_workbook
from sqlmodel import SQLModel

from app.core.config import settings
from app.database.database import async_session_factory, db_session, engine
from app.exceptions.exceptions import ExportExpiredError, ExportNotFoundError, PermissionDeniedError
from app.models.audit_log import AuditLog
from app.models.enums import ExportStatus, Status, StatusFilter
from app.models.platform_admin import PlatformAdmin
from app.schemas.export import AuditExportFilters, ExportCreate, UsersExportFilters
from app.services import export_service


@pytest.fixture()
async def engine_fixture():
    """Fresh schema on the app's own engine for every test (idempotent across runs)."""
    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.drop_all)
        await conn.run_sync(SQLModel.metadata.create_all)
    yield
    async with engine.begin() as conn:
        await conn.run_sync(SQLModel.metadata.drop_all)


async def _seed_audit_logs(n: int = 3) -> None:
    """Insert audit rows directly via the app's session factory."""
    async with async_session_factory() as session:
        for i in range(n):
            session.add(
                AuditLog(
                    actor=f"actor{i}@example.com",
                    actor_type="admin",
                    action="user.create",
                    resource_type="user",
                    resource_id=f"user-{i}",
                    details={"count": i},
                )
            )
        await session.commit()


async def _seed_platform_admins(n: int = 2) -> None:
    """Insert Platform Admin rows (the Users module) directly via the factory."""
    async with async_session_factory() as session:
        for i in range(n):
            session.add(
                PlatformAdmin(
                    email=f"seed{i}@example.com",
                    username=f"Seed Admin {i}",
                    hashed_password="x" * 60,
                    status=Status.ACTIVE if i % 2 == 0 else Status.INACTIVE,
                )
            )
        await session.commit()


async def test_full_export_lifecycle(engine_fixture, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))
    monkeypatch.setattr(export_service, "_spawn_generation", lambda _export_id: None)
    await _seed_audit_logs(3)

    async with db_session():
        created = await export_service.create_export(
            admin_email="admin@example.com",
            data=ExportCreate(
                module="audit", reason="Integration test", filters=AuditExportFilters()
            ),
        )
        assert created.status == ExportStatus.PENDING.value

        generated = await export_service._generate(created.id)
        assert generated.status == ExportStatus.READY.value
        # 3 seeded logs + the export.generated audit entry written by create_export
        assert generated.row_count == 4
        assert generated.expires_at is not None
        assert generated.file_path is not None

        path = Path(generated.file_path)
        assert path.is_file()

        workbook = load_workbook(path)
        assert workbook.sheetnames == ["Metadata", "Audit Events"]
        meta = dict(workbook["Metadata"].iter_rows(values_only=True))
        assert meta["Export ID"] == str(created.id)
        assert meta["Reason"] == "Integration test"
        assert meta["Classification"] == "Restricted"
        data = workbook["Audit Events"]
        data_rows = list(data.iter_rows(values_only=True))
        assert data_rows[0][0] == "Event ID"  # header row
        assert len(data_rows) == 5  # header + 3 seeded + 1 export.generated audit entry
        assert {row[2] for row in data_rows[1:]} == {  # actor column
            "actor0@example.com",
            "actor1@example.com",
            "actor2@example.com",
            "admin@example.com",
        }

        response = await export_service.download_export(
            export_id=created.id, admin_email="admin@example.com"
        )
        assert response.path == path
        assert response.media_type == export_service.XLSX_MEDIA_TYPE


async def test_export_ownership_is_enforced(engine_fixture, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))
    monkeypatch.setattr(export_service, "_spawn_generation", lambda _export_id: None)
    await _seed_audit_logs(1)

    async with db_session():
        created = await export_service.create_export(
            admin_email="owner@example.com",
            data=ExportCreate(module="audit", reason="Owner test", filters=AuditExportFilters()),
        )
        with pytest.raises(PermissionDeniedError):
            await export_service.get_export_status(
                export_id=created.id, admin_email="intruder@example.com"
            )
        with pytest.raises(PermissionDeniedError):
            await export_service.download_export(
                export_id=created.id, admin_email="intruder@example.com"
            )


async def test_export_expiry_blocks_download(engine_fixture, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))
    monkeypatch.setattr(export_service, "_spawn_generation", lambda _export_id: None)
    await _seed_audit_logs(1)

    async with db_session():
        created = await export_service.create_export(
            admin_email="admin@example.com",
            data=ExportCreate(module="audit", reason="Expiry test", filters=AuditExportFilters()),
        )
        created.expires_at = datetime.now(UTC).replace(tzinfo=None) - timedelta(minutes=1)
        with pytest.raises(ExportExpiredError):
            await export_service.download_export(
                export_id=created.id, admin_email="admin@example.com"
            )


async def test_export_not_found(engine_fixture) -> None:
    async with db_session():
        with pytest.raises(ExportNotFoundError):
            await export_service.get_export_status(
                export_id=uuid.uuid4(), admin_email="admin@example.com"
            )


async def test_users_export_lifecycle(engine_fixture, tmp_path, monkeypatch) -> None:
    """The users module: seed admins, export, verify workbook contents."""
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))
    monkeypatch.setattr(export_service, "_spawn_generation", lambda _export_id: None)
    await _seed_platform_admins(2)

    async with db_session():
        created = await export_service.create_export(
            admin_email="admin@example.com",
            data=ExportCreate(
                module="users", reason="User list for audit", filters=UsersExportFilters()
            ),
        )
        generated = await export_service._generate(created.id)

        assert generated.status == ExportStatus.READY.value
        assert generated.row_count == 2
        assert generated.classification == "Confidential"

        path = Path(generated.file_path or "")
        workbook = load_workbook(path)
        assert workbook.sheetnames == ["Metadata", "Users"]
        meta = dict(workbook["Metadata"].iter_rows(values_only=True))
        assert meta["Module"] == "Users"
        assert meta["Classification"] == "Confidential"

        data = workbook["Users"]
        rows = list(data.iter_rows(values_only=True))
        assert rows[0] == (
            "User ID",
            "Name",
            "Email",
            "Status",
            "Created At (UTC)",
            "Updated At (UTC)",
        )
        assert len(rows) == 3  # header + 2 admins
        assert {row[2] for row in rows[1:]} == {"seed0@example.com", "seed1@example.com"}


async def test_users_export_status_filter(engine_fixture, tmp_path, monkeypatch) -> None:
    """Users export with status=active keeps only active admins."""
    monkeypatch.setattr(settings, "export_dir", str(tmp_path))
    monkeypatch.setattr(export_service, "_spawn_generation", lambda _export_id: None)
    await _seed_platform_admins(2)  # one active, one inactive

    async with db_session():
        created = await export_service.create_export(
            admin_email="admin@example.com",
            data=ExportCreate(
                module="users",
                reason="Active users",
                filters=UsersExportFilters(status=StatusFilter.ACTIVE),
            ),
        )
        generated = await export_service._generate(created.id)
        assert generated.row_count == 1

        path = Path(generated.file_path or "")
        workbook = load_workbook(path)
        rows = list(workbook["Users"].iter_rows(values_only=True))
        assert len(rows) == 2  # header + 1 active admin
        assert rows[1][2] == "seed0@example.com"
        assert rows[1][3] == "active"
