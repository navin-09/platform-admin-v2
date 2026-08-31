"""xlsx writer tests (real files in tmp_path, read back with openpyxl)."""

from datetime import UTC, datetime

from openpyxl import load_workbook

from app.services import xlsx_writer


async def test_write_export_xlsx_creates_metadata_and_data_sheets(tmp_path) -> None:
    path = tmp_path / "export.xlsx"

    async def rows():
        yield ["1", datetime(2026, 8, 1, 10, 0, tzinfo=UTC), "admin@example.com", {"k": "v"}]
        yield ["2", None, "system", ["a", "b"]]

    count = await xlsx_writer.write_export_xlsx(
        path,
        metadata={"Reason": "Audit", "Classification": "Restricted", "Count": 2},
        metadata_sheet="Metadata",
        data_sheet="Events",
        headers=["ID", "Created", "Actor", "Details"],
        rows=rows(),
    )

    assert count == 2
    workbook = load_workbook(path)
    assert workbook.sheetnames == ["Metadata", "Events"]

    meta = workbook["Metadata"]
    meta_rows = list(meta.iter_rows(values_only=True))
    assert meta_rows[0] == ("Field", "Value")
    assert ("Reason", "Audit") in meta_rows
    assert ("Classification", "Restricted") in meta_rows

    data = workbook["Events"]
    data_rows = list(data.iter_rows(values_only=True))
    assert data_rows[0] == ("ID", "Created", "Actor", "Details")
    assert data_rows[1][0] == "1"
    assert data_rows[1][1] == "2026-08-01T10:00:00+00:00"  # datetime → ISO text
    assert data_rows[1][2] == "admin@example.com"
    assert data_rows[1][3] == '{"k": "v"}'  # dict → JSON text
    assert data_rows[2][1] in (None, "")  # None → empty cell
    assert data_rows[2][3] == '["a", "b"]'


async def test_write_export_xlsx_writes_zero_rows_with_headers_only(tmp_path) -> None:
    path = tmp_path / "empty.xlsx"

    async def rows():
        return
        yield  # pragma: no cover

    count = await xlsx_writer.write_export_xlsx(
        path,
        metadata={"Reason": "None"},
        metadata_sheet="Metadata",
        data_sheet="Events",
        headers=["A", "B"],
        rows=rows(),
    )
    assert count == 0
    workbook = load_workbook(path)
    data = workbook["Events"]
    assert list(data.iter_rows(values_only=True)) == [("A", "B")]


async def test_write_export_xlsx_replaces_existing_file(tmp_path) -> None:
    path = tmp_path / "export.xlsx"
    path.write_bytes(b"stale")

    async def rows():
        yield ["1"]

    count = await xlsx_writer.write_export_xlsx(
        path,
        metadata={"Reason": "R"},
        metadata_sheet="Metadata",
        data_sheet="Events",
        headers=["ID"],
        rows=rows(),
    )
    assert count == 1
    workbook = load_workbook(path)
    assert workbook["Events"]["A1"].value == "ID"
