"""Export service unit tests — repository and DB session fully mocked (see tests/unit/fakes.py)."""

import uuid
from contextlib import asynccontextmanager
from datetime import UTC, datetime, timedelta
from pathlib import Path
from types import SimpleNamespace
from typing import Any
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook

from app.exceptions.exceptions import (
    ExportExpiredError,
    ExportNotFoundError,
    ExportTooLargeError,
    PermissionDeniedError,
    ValidationError,
)
from app.models.enums import ExportStatus
from app.models.export import Export
from app.schemas.export import AuditExportFilters, ExportCreate, UsersExportFilters
from app.services import export_service

# --------------------------------------------------------------------------- #
# Test doubles
# --------------------------------------------------------------------------- #


class FakeExportRepository:
    """In-memory stand-in for ``app.repositories.export_repository``."""

    def __init__(self) -> None:
        self.export: Export | None = None  # returned by get_export
        self.row_count = 0  # returned by count_*
        self.streamed: list[Any] = []  # rows yielded by stream_*
        self.created: Export | None = None
        self.created_kwargs: dict[str, Any] = {}
        self.updated: list[Export] = []
        self.audit_filters_calls: list[tuple[Any, ...]] = []
        self.user_filters_calls: list[tuple[Any, ...]] = []
        self.spawned: list[uuid.UUID] = []
        self.audit_record = AsyncMock()

    async def create_export(self, **kwargs: Any) -> Export:
        self.created_kwargs = kwargs
        self.created = Export(**kwargs)
        return self.created

    async def get_export(self, export_id: uuid.UUID) -> Export | None:
        return self.export

    async def update_export(self, export: Export) -> None:
        self.updated.append(export)

    async def count_audit_logs(self, **kwargs: Any) -> int:
        self.audit_filters_calls.append(tuple(kwargs.values()))
        return self.row_count

    async def count_platform_admins(self, **kwargs: Any) -> int:
        self.user_filters_calls.append(tuple(kwargs.values()))
        return self.row_count

    async def stream_audit_logs(self, **kwargs: Any):
        for row in self.streamed:
            yield row

    async def stream_platform_admins(self, **kwargs: Any):
        for row in self.streamed:
            yield row


@pytest.fixture()
def fake_repo(monkeypatch) -> FakeExportRepository:
    """Wire the fake repository + stubs into the service under test."""
    repo = FakeExportRepository()

    @asynccontextmanager
    async def _noop_db_session():
        yield

    def _spawn(export_id: uuid.UUID) -> None:
        repo.spawned.append(export_id)

    monkeypatch.setattr(export_service, "export_repository", repo)
    monkeypatch.setattr(export_service, "db_session", _noop_db_session)
    monkeypatch.setattr(export_service, "_spawn_generation", _spawn)
    monkeypatch.setattr(export_service.audit_service, "record", repo.audit_record)
    return repo


def _export(
    *,
    status: str = ExportStatus.READY.value,
    file_path: str | None = "exports/sample.xlsx",
    expires_at: datetime | None = None,
    created_by: str = "admin@example.com",
    module: str = "audit",
    filters: dict[str, object] | None = None,
) -> Export:
    return Export(
        id=uuid.uuid4(),
        module=module,
        reason="Regulatory request",
        file_format="xlsx",
        classification="Restricted",
        filters=filters or {"actor": None, "action": "All"},
        status=status,
        row_count=2,
        file_path=file_path,
        expires_at=expires_at or (_utcnow() + timedelta(hours=1)),
        created_by=created_by,
    )


def _utcnow() -> datetime:
    """Naive UTC now — matches the service's utcnow() (Postgres timestamps)."""
    return datetime.now(UTC).replace(tzinfo=None)


def _audit_row() -> SimpleNamespace:
    """One fake AuditLog-shaped row matching the audit export spec columns."""
    return SimpleNamespace(
        id=uuid.uuid4(),
        created_at=_utcnow(),
        actor="admin@example.com",
        actor_type="admin",
        action="user.create",
        resource_type="user",
        resource_id="user-1",
        details={"k": "v"},
        payload=None,
        response=None,
        url="/api/v1/users",
        request_id="req-1",
        ip_address="203.0.113.7",
        user_agent="pytest",
    )


# --------------------------------------------------------------------------- #
# create_export
# --------------------------------------------------------------------------- #


async def test_create_export_success(fake_repo) -> None:
    fake_repo.row_count = 5

    export = await export_service.create_export(
        admin_email="admin@example.com",
        data=ExportCreate(module="audit", reason="Quarterly review"),
    )

    assert export.status == ExportStatus.PENDING.value
    assert fake_repo.created_kwargs["module"] == "audit"
    assert fake_repo.created_kwargs["created_by"] == "admin@example.com"
    assert fake_repo.created_kwargs["reason"] == "Quarterly review"
    assert fake_repo.created_kwargs["classification"] == "Restricted"
    assert fake_repo.spawned == [export.id]
    action = fake_repo.audit_record.await_args.kwargs["action"]
    assert action == "export.generated"
    assert fake_repo.audit_record.await_args.kwargs["resource_id"] == str(export.id)


async def test_create_export_users_module_uses_users_filters(fake_repo) -> None:
    fake_repo.row_count = 1

    await export_service.create_export(
        admin_email="admin@example.com",
        data=ExportCreate(
            module="users",
            reason="User list",
            filters=UsersExportFilters(status="active"),
        ),
    )

    assert fake_repo.created_kwargs["classification"] == "Confidential"
    assert fake_repo.user_filters_calls == [(None, "active")]


async def test_create_export_failure_over_max_rows(fake_repo) -> None:
    fake_repo.row_count = 100_001

    with pytest.raises(ExportTooLargeError):
        await export_service.create_export(
            admin_email="admin@example.com",
            data=ExportCreate(module="audit", reason="Too big"),
        )
    assert fake_repo.spawned == []  # nothing spawned, nothing created


async def test_create_export_failure_unsupported_module(fake_repo) -> None:
    """The spec registry rejects unknown modules (unreachable via the Literal)."""
    with pytest.raises(ValidationError):
        export_service._spec("bogus")
    assert fake_repo.spawned == []


async def test_create_export_failure_wrong_filter_shape(fake_repo) -> None:
    """Audit-shaped filters on the users module must be rejected, never accepted."""
    data = ExportCreate.model_validate(
        {"module": "users", "reason": "r", "filters": {"actor": "x"}}
    )
    with pytest.raises(ValidationError):
        await export_service.create_export(admin_email="admin@example.com", data=data)
    assert fake_repo.spawned == []


# --------------------------------------------------------------------------- #
# get_export_status
# --------------------------------------------------------------------------- #


async def test_get_export_status_success(fake_repo) -> None:
    export = _export()
    fake_repo.export = export

    result = await export_service.get_export_status(
        export_id=export.id, admin_email="admin@example.com"
    )
    assert result is export


async def test_get_export_status_failure_not_found(fake_repo) -> None:
    fake_repo.export = None

    with pytest.raises(ExportNotFoundError):
        await export_service.get_export_status(
            export_id=uuid.uuid4(), admin_email="admin@example.com"
        )


async def test_get_export_status_failure_not_owner(fake_repo) -> None:
    fake_repo.export = _export(created_by="other@example.com")

    with pytest.raises(PermissionDeniedError):
        await export_service.get_export_status(
            export_id=fake_repo.export.id, admin_email="admin@example.com"
        )


# --------------------------------------------------------------------------- #
# download_export
# --------------------------------------------------------------------------- #


async def test_download_export_success(fake_repo, tmp_path) -> None:
    target = tmp_path / "real.xlsx"
    target.write_bytes(b"PK fake xlsx")
    export = _export(file_path=str(target))
    fake_repo.export = export

    response = await export_service.download_export(
        export_id=export.id, admin_email="admin@example.com"
    )

    assert response.path == target
    assert response.media_type == export_service.XLSX_MEDIA_TYPE
    action = fake_repo.audit_record.await_args.kwargs["action"]
    assert action == "export.downloaded"
    assert fake_repo.audit_record.await_args.kwargs["resource_id"] == str(export.id)


async def test_download_export_success_regenerates_when_file_missing(
    fake_repo, tmp_path, monkeypatch
) -> None:
    export = _export(file_path=str(tmp_path / "missing.xlsx"))
    fake_repo.export = export
    regenerated_path = tmp_path / "regenerated.xlsx"
    regenerated_path.write_bytes(b"PK fake xlsx")
    regenerated = _export(file_path=str(regenerated_path))
    monkeypatch.setattr(export_service, "_generate", AsyncMock(return_value=regenerated))

    response = await export_service.download_export(
        export_id=export.id, admin_email="admin@example.com"
    )

    assert response.path == regenerated_path
    export_service._generate.assert_awaited_once_with(export.id)


async def test_download_export_failure_expired(fake_repo) -> None:
    export = _export(expires_at=_utcnow() - timedelta(minutes=1))
    fake_repo.export = export

    with pytest.raises(ExportExpiredError):
        await export_service.download_export(export_id=export.id, admin_email="admin@example.com")
    assert export.status == ExportStatus.EXPIRED.value
    assert fake_repo.updated == [export]


async def test_download_export_failure_not_found(fake_repo) -> None:
    fake_repo.export = None

    with pytest.raises(ExportNotFoundError):
        await export_service.download_export(
            export_id=uuid.uuid4(), admin_email="admin@example.com"
        )


async def test_download_export_failure_not_owner(fake_repo, tmp_path) -> None:
    target = tmp_path / "real.xlsx"
    target.write_bytes(b"PK fake xlsx")
    fake_repo.export = _export(file_path=str(target), created_by="other@example.com")

    with pytest.raises(PermissionDeniedError):
        await export_service.download_export(
            export_id=fake_repo.export.id, admin_email="admin@example.com"
        )


# --------------------------------------------------------------------------- #
# _generate (file generation)
# --------------------------------------------------------------------------- #


async def test_generate_success_writes_real_xlsx(fake_repo, tmp_path, monkeypatch) -> None:
    monkeypatch.setattr("app.core.config.settings.export_dir", str(tmp_path))
    export = _export(status=ExportStatus.PENDING.value, file_path=None)
    fake_repo.export = export
    fake_repo.streamed = [_audit_row(), _audit_row()]

    generated = await export_service._generate(export.id)

    assert generated.status == ExportStatus.READY.value
    assert generated.row_count == 2
    assert generated.file_path is not None
    assert generated.expires_at is not None

    workbook = load_workbook(Path(generated.file_path))
    assert workbook.sheetnames == ["Metadata", "Audit Events"]
    meta = dict(workbook["Metadata"].iter_rows(values_only=True))
    assert meta["Export ID"] == str(export.id)
    assert meta["Classification"] == "Restricted"
    data = list(workbook["Audit Events"].iter_rows(values_only=True))
    assert data[0][0] == "Event ID"
    assert len(data) == 3  # header + 2 rows


async def test_generate_failure_export_missing(fake_repo) -> None:
    fake_repo.export = None

    with pytest.raises(ExportNotFoundError):
        await export_service._generate(uuid.uuid4())


async def test_generate_task_failure_marks_failed(fake_repo, monkeypatch) -> None:
    export = _export(status=ExportStatus.PENDING.value, file_path=None)
    fake_repo.export = export

    async def _boom(export_id: uuid.UUID) -> Export:
        raise RuntimeError("db down")

    monkeypatch.setattr(export_service, "_generate", _boom)

    await export_service._generate_task(export.id)

    assert export.status == ExportStatus.FAILED.value
    assert export.generation_error is not None
    assert fake_repo.updated == [export]


# --------------------------------------------------------------------------- #
# filters → repository dispatch
# --------------------------------------------------------------------------- #


async def test_create_export_passes_resolved_audit_filters(fake_repo) -> None:
    fake_repo.row_count = 3

    await export_service.create_export(
        admin_email="admin@example.com",
        data=ExportCreate(
            module="audit",
            reason="r",
            filters=AuditExportFilters(actor="admin@example.com", action="user.create"),
        ),
    )

    assert fake_repo.audit_filters_calls == [("admin@example.com", "user.create", None, None)]
